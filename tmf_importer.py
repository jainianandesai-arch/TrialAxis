"""
eTMF Status Importer
====================
Reads the eTMF Excel export and maps TMF document statuses
into trial_data.py.

Composite key: TAX Study ID + Document Type
- TAX Study ID comes from their eTMF system via Excel — never generated here
- CRA-only fields (visit notes, action items, knowledge) are NEVER overwritten

Usage:
    python tmf_importer.py
    python tmf_importer.py data/exports/eTMF_Status_Report.xlsx
    python tmf_importer.py data/exports/eTMF_Status_Report.xlsx --dry-run

In production:
    Power Automate detects new export → triggers this script automatically
"""

import sys
import re
import os
from pathlib import Path
from datetime import datetime, date

# ── Paths ──────────────────────────────────────────────────────────────────────
DEFAULT_EXPORT  = Path("data/exports/eTMF_Status_Report.xlsx")
TRIAL_DATA_CANDIDATES = [Path("trial_data.py"), Path("src/trial_data.py")]

# ── Column names in Excel ──────────────────────────────────────────────────────
COL_TAX_ID      = "TAX Study ID"
COL_DOC_TYPE    = "Document Type"
COL_STATUS      = "Status"
COL_DATE_FILED  = "Date Filed"
COL_VERSION     = "Version"
COL_EXPIRY_DATE = "Expiry Date"
COL_DAYS_EXPIRY = "Days to Expiry"
COL_RESP_PARTY  = "Responsible Party"
COL_NOTES       = "Notes"
COL_EXPORT_TS   = "Export Timestamp"

# CRA-only fields — NEVER overwritten by this importer
CRA_ONLY_FIELDS = {
    "cra_visit_date", "cra_visit_notes",
    "action_items",   "institutional_knowledge"
}

def find_trial_data() -> Path:
    for p in TRIAL_DATA_CANDIDATES:
        if p.exists():
            return p
    raise FileNotFoundError("trial_data.py not found.")

def read_export(filepath: Path) -> list:
    """Read TMF Status Report sheet from Excel or CSV."""
    if not filepath.exists():
        raise FileNotFoundError(f"Export not found: {filepath}")

    if filepath.suffix == ".xlsx":
        import openpyxl
        wb = openpyxl.load_workbook(str(filepath), data_only=True)

        # Known format — header rows per sheet (rows above are Veeva export banners)
        SHEET_HEADER_ROWS = {
            "Study Registry":    3,
            "TMF Status Report": 3,
            "Import Mapping":    2,
        }

        # Use TMF Status Report sheet
        ws = None
        for name in ["TMF Status Report", wb.sheetnames[0]]:
            if name in wb.sheetnames:
                ws = wb[name]
                break

        header_row = SHEET_HEADER_ROWS.get(ws.title, 3)

        # Verify — fallback scan if needed
        actual_val = str(ws.cell(row=header_row, column=1).value or "").strip()
        if actual_val != "TAX Study ID":
            for i in range(1, 10):
                if str(ws.cell(row=i, column=1).value or "").strip() == "TAX Study ID":
                    header_row = i
                    break

        headers = [ws.cell(row=header_row, column=c).value
                   for c in range(1, ws.max_column + 1)]
        rows = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if row[0]:
                rows.append(dict(zip(headers, row)))
        return rows

    elif filepath.suffix == ".csv":
        import csv
        with open(filepath, newline="", encoding="utf-8-sig") as f:
            return list(csv.DictReader(f))

    raise ValueError(f"Unsupported format: {filepath.suffix}")

def normalize_status(status, days_to_expiry=None) -> str:
    """Map eTMF status string → internal status. Auto-expire on negative days."""
    if not status:
        return "Missing"
    s = str(status).strip().lower()
    if days_to_expiry is not None:
        try:
            if int(days_to_expiry) < 0 and s == "complete":
                return "Expired"
        except (ValueError, TypeError):
            pass
    return {
        "complete":     "Complete",
        "missing":      "Missing",
        "expired":      "Expired",
        "needs review": "Needs Review",
        "in progress":  "Needs Review",
        "pending":      "Needs Review",
        "not required": "Complete",
    }.get(s, str(status).strip())

def normalize_date(val) -> str:
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    return s if s else None

def run_import(filepath: Path, dry_run: bool = False) -> dict:
    import pprint as _pprint

    print("\n" + "=" * 60)
    print("  eTMF STATUS IMPORTER -- TMF Intelligence System")
    print(f"  File:  {filepath.name}")
    print(f"  Key:   TAX Study ID + Document Type")
    print(f"  Mode:  {'DRY RUN' if dry_run else 'LIVE UPDATE'}")
    print("=" * 60)

    # Read export
    print(f"\n  Reading export...")
    rows = read_export(filepath)
    print(f"  OK {len(rows)} document records")

    # Group by TAX Study ID
    export_index = {}
    for row in rows:
        tax = str(row.get(COL_TAX_ID) or "").strip()
        doc = str(row.get(COL_DOC_TYPE) or "").strip()
        if tax and doc:
            export_index.setdefault(tax, {})[doc] = row

    print(f"  OK {len(export_index)} studies in export")

    # Load trial_data.py by executing it — avoids fragile regex on Python source
    td_path = find_trial_data()
    source  = td_path.read_text(encoding="utf-8")
    ns = {}
    exec(compile(source, str(td_path), "exec"), ns)
    trials = ns["TRIALS"]
    print(f"  OK trial_data.py loaded ({len(trials)} studies)")

    changes   = []
    no_match  = []
    unchanged = 0

    print(f"\n  Processing...")

    for tax_id, docs in export_index.items():
        if tax_id not in trials:
            print(f"  !  {tax_id} -- not in TRIALS, skipping")
            continue

        for doc_name, row in docs.items():
            new_status  = normalize_status(
                row.get(COL_STATUS), row.get(COL_DAYS_EXPIRY)
            )
            new_date    = normalize_date(row.get(COL_DATE_FILED))
            new_version = str(row.get(COL_VERSION) or "").strip() or None

            tmf_docs = trials[tax_id].get("tmf_documents", {})
            if doc_name not in tmf_docs:
                no_match.append(f"{tax_id} | {doc_name}")
                continue

            current     = tmf_docs[doc_name]
            old_status  = current.get("status")
            old_date    = current.get("date")
            old_version = current.get("version")

            if old_status == new_status and old_date == new_date and old_version == new_version:
                unchanged += 1
                continue

            # Update in memory (CRA-only fields are untouched — not in tmf_documents)
            tmf_docs[doc_name]["status"]  = new_status
            tmf_docs[doc_name]["date"]    = new_date
            tmf_docs[doc_name]["version"] = new_version

            if old_status != new_status:
                print(f"  >>  {tax_id} | {doc_name}: {old_status} -> {new_status}")
            else:
                print(f"  >>  {tax_id} | {doc_name}: date/version updated ({new_status})")

            changes.append({
                "tax_id": tax_id, "document": doc_name,
                "old": old_status, "new": new_status
            })

    # Write back using pprint — round-trips cleanly through Python's own parser
    if not dry_run and changes:
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        sync_line = f'# Last eTMF sync: {ts}\n'

        # Preserve header (module docstring + any comments) before TRIALS =
        marker = "TRIALS = "
        marker_idx = source.find(marker)
        header = source[:marker_idx] if marker_idx != -1 else ""
        if header.startswith("# Last eTMF sync:"):
            header = re.sub(r'^# Last eTMF sync:.*\n', sync_line, header)
        else:
            header = sync_line + header

        new_source = (
            header
            + "TRIALS = "
            + _pprint.pformat(trials, indent=4, sort_dicts=False)
            + "\n"
        )

        with open(td_path, "w", encoding="utf-8") as f:
            f.write(new_source)
        os.utime(td_path, None)
        print(f"\n  OK trial_data.py updated -- Streamlit will reload")

    # Summary
    print("\n" + "=" * 60)
    print("  IMPORT COMPLETE")
    print(f"  Status changes : {len(changes)}")
    print(f"  Unchanged      : {unchanged}")
    if no_match:
        print(f"  No match       : {len(no_match)}")
        for nm in no_match:
            print(f"    ! {nm}")
    if dry_run:
        print("  DRY RUN -- nothing written")
    print("=" * 60)

    return {"changes": changes, "unchanged": unchanged, "no_match": no_match}

if __name__ == "__main__":
    filepath = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_EXPORT
    dry_run  = "--dry-run" in sys.argv
    try:
        run_import(filepath, dry_run=dry_run)
    except Exception as e:
        print(f"\nFAIL Import failed: {e}")
        import traceback; traceback.print_exc()
        sys.exit(1)
