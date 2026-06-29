"""
Bulk setup -- runs TMFOrchestrator on all PDFs in data/pdfs/
No separate chunking/embedding logic. Everything goes through agents.py.

Usage:
    python setup_data.py           # skip already-indexed studies
    python setup_data.py --force   # re-index everything
    python setup_data.py --clean   # wipe ChromaDB + JSON, then re-index all
"""

import os
import sys
import shutil
from pathlib import Path
from dotenv import load_dotenv

load_dotenv(dotenv_path=Path(__file__).parent / ".env", override=True)

_key = os.environ.get("ANTHROPIC_API_KEY", "")
print(f"  API key: {'loaded (' + _key[:12] + '...)' if _key else 'NOT FOUND'}")

PDF_DIR     = Path("data/pdfs")
CHROMA_DIR  = Path("data/chroma_db")
INDEX_FILE  = Path("data/trial_index.json")
EXPORT_FILE = Path("data/exports/eTMF_Status_Report.xlsx")

FORCE = "--force" in sys.argv
CLEAN = "--clean" in sys.argv

PDF_DIR.mkdir(parents=True, exist_ok=True)
CHROMA_DIR.mkdir(parents=True, exist_ok=True)


class FileWrapper:
    def __init__(self, path):
        self.name = Path(path).name
        self._path = Path(path)

    def read(self):
        return self._path.read_bytes()


def load_study_registry() -> list:
    if not EXPORT_FILE.exists():
        print(f"  FAIL Export file not found: {EXPORT_FILE}")
        return []

    import openpyxl
    wb = openpyxl.load_workbook(str(EXPORT_FILE), data_only=True)

    ws = None
    for name in ["Study Registry", "Portfolio Summary"]:
        if name in wb.sheetnames:
            ws = wb[name]
            break
    if not ws:
        print("  FAIL 'Study Registry' sheet not found in Excel")
        return []

    HEADER_ROW = 3
    actual_val = str(ws.cell(row=HEADER_ROW, column=1).value or "").strip()
    if actual_val != "TAX Study ID":
        for i in range(1, 10):
            if str(ws.cell(row=i, column=1).value or "").strip() == "TAX Study ID":
                HEADER_ROW = i
                break

    if str(ws.cell(row=HEADER_ROW, column=1).value or "").strip() != "TAX Study ID":
        raise ValueError("'TAX Study ID' header not found in Study Registry sheet")

    headers = [ws.cell(row=HEADER_ROW, column=c).value
               for c in range(1, ws.max_column + 1)]
    studies = []
    for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        if not row[0] or not str(row[0]).startswith("TAX"):
            continue
        studies.append(dict(zip(headers, row)))
    return studies


def _clean_msg(message: str) -> str:
    """Strip markdown and non-ASCII for Windows console output."""
    replacements = [
        ("**", ""), ("✓", "OK"), ("✗", "FAIL"), ("⚠", "!"),
        ("❌", "FAIL"), ("\U0001f4e5", "[Upload]"), ("\U0001f50d", "[Meta]"),
        ("\U0001f4cb", "[Registry]"), ("\U0001f4da", "[Index]"), ("\U0001f504", "[Sync]"),
        ("\U0001f6a8", "ALERT"), ("\U0001f389", ">>"), ("—", "--"), ("→", "->"),
    ]
    for old, new in replacements:
        message = message.replace(old, new)
    return message.encode("ascii", errors="replace").decode("ascii")


def main():
    print("\n" + "=" * 60)
    print("  TMF INTELLIGENCE SYSTEM -- BULK SETUP")
    print("  Model: Claude Sonnet 4.6")
    print("  Key:   TAX Study ID from Excel Study Registry")
    print("=" * 60)

    if not os.environ.get("ANTHROPIC_API_KEY"):
        print("\nFAIL ANTHROPIC_API_KEY not set.")
        return

    import chromadb
    sys.path.insert(0, str(Path(__file__).parent))
    from agents import TMFOrchestrator

    # --clean: wipe ChromaDB and JSON index before rebuilding
    if CLEAN:
        print("\n  --clean flag detected -- wiping existing index...")
        if CHROMA_DIR.exists():
            try:
                shutil.rmtree(str(CHROMA_DIR))
                CHROMA_DIR.mkdir(parents=True, exist_ok=True)
                print("  OK ChromaDB cleared")
            except Exception:
                print("  ! Cannot delete chroma_db folder (Windows lock) -- clearing files instead")
                for item in CHROMA_DIR.rglob("*"):
                    try:
                        if item.is_file():
                            item.unlink()
                    except Exception:
                        pass
                CHROMA_DIR.mkdir(parents=True, exist_ok=True)
                print("  OK ChromaDB contents cleared")
        if INDEX_FILE.exists():
            INDEX_FILE.unlink()
            print("  OK trial_index.json cleared")

    # ChromaDB client for pre-flight skip check
    chroma_cli = chromadb.PersistentClient(path=str(CHROMA_DIR))
    collection = chroma_cli.get_or_create_collection(
        name="trial_chunks",
        metadata={"hnsw:space": "cosine"}
    )

    # Load Study Registry from Excel
    print(f"\n  Loading Study Registry from {EXPORT_FILE}...")
    studies = load_study_registry()
    if not studies:
        print("  FAIL No studies found. Check the Excel file.")
        return
    print(f"  OK {len(studies)} studies in registry")

    print("\n" + "=" * 60)
    print("  PROCESSING STUDIES")
    print("=" * 60)

    processed = 0
    skipped   = 0
    failed    = 0

    for study in studies:
        tax_id = str(study.get("TAX Study ID") or "").strip()
        if not tax_id.startswith("TAX"):
            continue

        # Skip if already in ChromaDB and not --force / --clean
        if not FORCE and not CLEAN:
            try:
                existing = collection.get(where={"tax_id": tax_id})["ids"]
                if existing:
                    print(f"\n  {tax_id} -- already indexed ({len(existing)} chunks), skipping")
                    skipped += 1
                    continue
            except Exception:
                pass

        # Find the PDF
        pdf_path = None
        pdf_filename = str(study.get("PDF Filename") or "").strip()
        if pdf_filename:
            p = PDF_DIR / pdf_filename
            if p.exists():
                pdf_path = p
        if not pdf_path:
            matches = list(PDF_DIR.glob(f"{tax_id}*.pdf"))
            if matches:
                pdf_path = matches[0]

        if not pdf_path:
            print(f"\n  {tax_id} -- No PDF found, skipping")
            print(f"    Expected: {pdf_filename or tax_id + '_*.pdf'}")
            failed += 1
            continue

        study_name = str(study.get("Study Name") or "").strip()
        print(f"\n  {tax_id} -- {study_name} [{pdf_path.name}]")

        # Run TMFOrchestrator -- same pipeline as ingest_protocol.py
        orchestrator = TMFOrchestrator()
        generator = orchestrator.run(
            uploaded_file=FileWrapper(pdf_path),
            existing_trials={}
        )

        ok = False
        for message in generator:
            if message == "__SUCCESS__":
                ok = True
                continue
            if message == "__UNREGISTERED__":
                print("    ! Not in registry -- moved to unregistered/")
                break
            print(f"    {_clean_msg(message)}")

        if ok:
            processed += 1
        else:
            failed += 1

    # Final summary
    print("\n" + "=" * 60)
    print("  SETUP COMPLETE")
    print("=" * 60)
    print(f"  Processed: {processed} | Skipped: {skipped} | Failed: {failed}")

    # Reload collection for final chunk count
    try:
        chroma_cli2 = chromadb.PersistentClient(path=str(CHROMA_DIR))
        col2 = chroma_cli2.get_or_create_collection(
            name="trial_chunks",
            metadata={"hnsw:space": "cosine"}
        )
        all_ids = col2.get()["ids"]
        by_tax = {}
        for chunk_id in all_ids:
            parts = chunk_id.rsplit("_", 1)
            tid = parts[0] if len(parts) == 2 else chunk_id
            by_tax[tid] = by_tax.get(tid, 0) + 1
        print("\n  Indexed studies:")
        for tid, count in sorted(by_tax.items()):
            print(f"    OK {tid} -- {count} chunks")
        print(f"  Total chunks: {len(all_ids)}")
    except Exception as e:
        print(f"  (Could not read final chunk count: {e})")

    print(f"\n  ChromaDB: {CHROMA_DIR}/")
    print(f"  JSON:     {INDEX_FILE}")
    print("\n  Next: python tmf_importer.py -> streamlit run app.py")
    print("=" * 60)


if __name__ == "__main__":
    main()
